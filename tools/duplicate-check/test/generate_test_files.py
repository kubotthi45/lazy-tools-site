"""
重複データ検出ツールのテスト用Excelファイル生成スクリプト
5つのファイルを生成する
"""
import openpyxl
from pathlib import Path
from datetime import datetime

OUT_DIR = Path(__file__).parent


def create_test1():
    """テスト1: 基本パターン — 完全一致の重複"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "名簿"

    ws.append(["名前", "メールアドレス", "住所", "電話番号"])
    ws.append(["田中太郎", "tanaka@example.com", "東京都新宿区1-1-1", "03-1111-2222"])
    ws.append(["佐藤花子", "sato@example.com", "大阪府大阪市北区2-2-2", "06-3333-4444"])
    ws.append(["田中太郎", "tanaka@example.com", "東京都新宿区1-1-1", "03-1111-2222"])  # 完全重複
    ws.append(["鈴木一郎", "suzuki@example.com", "愛知県名古屋市中区3-3-3", "052-5555-6666"])
    ws.append(["佐藤花子", "sato@example.com", "大阪府大阪市北区2-2-2", "06-3333-4444"])  # 完全重複
    ws.append(["高橋美咲", "takahashi@example.com", "福岡県福岡市博多区4-4-4", "092-7777-8888"])
    ws.append(["田中太郎", "tanaka@example.com", "東京都新宿区1-1-1", "03-1111-2222"])  # 3回目の重複
    ws.append(["渡辺健", "watanabe@example.com", "北海道札幌市中央区5-5-5", "011-9999-0000"])

    wb.save(OUT_DIR / "test1_基本重複.xlsx")
    print("  test1_基本重複.xlsx 作成完了")


def create_test2():
    """テスト2: 同姓同名だが別人 — 列の組み合わせが重要"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "顧客リスト"

    ws.append(["名前", "メールアドレス", "住所", "生年月日"])
    ws.append(["田中太郎", "tanaka1@example.com", "東京都新宿区1-1-1", "1990/4/1"])     # 田中A
    ws.append(["田中太郎", "tanaka2@example.com", "大阪府大阪市北区2-2-2", "1985/8/15"])  # 田中B（別人）
    ws.append(["田中太郎", "tanaka1@example.com", "東京都新宿区1-1-1", "1990/4/1"])     # 田中Aの重複
    ws.append(["佐藤花子", "sato@example.com", "福岡県福岡市5-5-5", "1992/12/25"])
    ws.append(["鈴木一郎", "suzuki@example.com", "愛知県名古屋市3-3-3", "1988/3/10"])
    # 同じ住所だが別人
    ws.append(["山田次郎", "yamada@example.com", "東京都新宿区1-1-1", "1995/7/20"])    # 田中Aと住所が同じ
    ws.append(["佐藤花子", "sato@example.com", "福岡県福岡市5-5-5", "1992/12/25"])     # 佐藤の重複

    wb.save(OUT_DIR / "test2_同姓同名.xlsx")
    print("  test2_同姓同名.xlsx 作成完了")


def create_test3():
    """テスト3: 全角半角・大文字小文字・空白の罠"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "データ整形テスト"

    ws.append(["商品コード", "商品名", "価格", "カテゴリ"])
    # 全角半角の違い
    ws.append(["ABC-001", "ノートPC", 98000, "電子機器"])
    ws.append(["ＡＢＣ−００１", "ノートPC", 98000, "電子機器"])  # 全角版（同じものとみなすべきか？）
    # 大文字小文字の違い
    ws.append(["def-002", "マウス", 3500, "周辺機器"])
    ws.append(["DEF-002", "マウス", 3500, "周辺機器"])  # 大文字版
    # 前後の空白
    ws.append(["GHI-003", "キーボード", 5000, "周辺機器"])
    ws.append(["GHI-003 ", "キーボード", 5000, "周辺機器"])  # 末尾に空白
    ws.append([" GHI-003", "キーボード", 5000, "周辺機器"])  # 先頭に空白
    # 全角スペース
    ws.append(["JKL-004", "モニター", 35000, "電子機器"])
    ws.append(["JKL-004", "モニター　", 35000, "電子機器"])  # 全角スペースが末尾に
    # 完全に別物
    ws.append(["MNO-005", "USBケーブル", 800, "周辺機器"])

    wb.save(OUT_DIR / "test3_全角半角空白.xlsx")
    print("  test3_全角半角空白.xlsx 作成完了")


def create_test4():
    """テスト4: 数式・小数点・数値の丸め誤差"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上データ"

    ws.append(["ID", "商品", "単価", "数量", "合計", "税込"])
    # 通常の重複
    ws.append([1, "商品A", 1000, 5, 5000, 5500])
    ws.append([2, "商品B", 2000, 3, 6000, 6600])
    ws.append([1, "商品A", 1000, 5, 5000, 5500])  # IDと内容が完全重複

    # 数式で計算した合計（値は同じだが数式で入力）
    ws["E5"] = 1000  # 単価
    ws["F5"] = 5  # 数量
    ws.append([3, "商品C", None, None, None, None])
    ws["C5"] = 1000
    ws["D5"] = 5
    ws["E5"] = "=C5*D5"  # =5000
    ws["F5"] = "=E5*1.1"  # =5500

    # 浮動小数点の誤差が出るケース
    ws.append([4, "商品D", 100.1, 3, None, None])
    ws["E7"] = "=C7*D7"  # =300.30000000000004 (浮動小数点誤差)

    ws.append([5, "商品D", 100.1, 3, 300.3, None])  # 手入力の300.3（上と「同じ」はずだが誤差あり）
    ws["F8"] = "=E8*1.1"

    # 小数点の桁数の違い
    ws.append([6, "商品E", 99.999, 1, 99.999, None])
    ws.append([7, "商品E", 100.0, 1, 100.0, None])   # 四捨五入すれば同じ
    ws.append([8, "商品E", 99.999, 1, 99.999, None])  # 6と完全一致

    # 整数と小数
    ws.append([9, "商品F", 500, 2, 1000, None])
    ws.append([10, "商品F", 500.0, 2, 1000.0, None])  # 500 vs 500.0

    wb.save(OUT_DIR / "test4_数式と丸め誤差.xlsx")
    print("  test4_数式と丸め誤差.xlsx 作成完了")


def create_test5():
    """テスト5: 空セル・混在データ型・大量データ"""
    wb = openpyxl.Workbook()

    # シート1: 空セルが混在
    ws1 = wb.active
    ws1.title = "空セル混在"
    ws1.append(["名前", "部署", "電話番号", "備考"])
    ws1.append(["田中", "営業部", "03-1111-2222", ""])
    ws1.append(["佐藤", "", "06-3333-4444", "新入社員"])       # 部署が空
    ws1.append(["田中", "営業部", "03-1111-2222", ""])          # 重複
    ws1.append(["鈴木", "開発部", "", ""])                      # 電話番号と備考が空
    ws1.append(["佐藤", "", "06-3333-4444", "新入社員"])       # 重複（部署が空のまま）
    ws1.append(["", "", "", ""])                                 # 全て空の行
    ws1.append(["", "", "", ""])                                 # 全て空の行（重複？）
    ws1.append(["高橋", "経理部", "052-5555-6666", None])

    # シート2: データ型が混在
    ws2 = wb.create_sheet("データ型混在")
    ws2.append(["ID", "値", "日付", "フラグ"])
    ws2.append([1, "100", datetime(2026, 4, 1), True])         # 文字列の"100"
    ws2.append([2, 100, datetime(2026, 4, 1), True])           # 数値の100（文字列の"100"と同じ？）
    ws2.append([3, "100", datetime(2026, 4, 1), True])         # 1行目と同じ
    ws2.append([4, "abc", datetime(2026, 4, 2), False])
    ws2.append([5, "ABC", datetime(2026, 4, 2), False])        # 大文字小文字

    # シート3: 大量データ（100行）
    ws3 = wb.create_sheet("大量データ")
    ws3.append(["連番", "グループ", "値"])
    import random
    random.seed(42)
    groups = ["A", "B", "C", "D", "E"]
    for i in range(1, 101):
        g = random.choice(groups)
        v = random.randint(1, 20)
        ws3.append([i, g, v])
    # 意図的に重複行を挿入
    ws3.append([101, "A", 5])   # 上のどこかにA-5があるはず
    ws3.append([102, "B", 10])  # 上のどこかにB-10があるはず

    wb.save(OUT_DIR / "test5_空セル型混在大量.xlsx")
    print("  test5_空セル型混在大量.xlsx 作成完了")


def create_test6():
    """テスト6: ヘッダーが1行目にないパターン"""
    wb = openpyxl.Workbook()

    # シート1: 1〜2行目がタイトル・日付、3行目がヘッダー
    ws1 = wb.active
    ws1.title = "3行目ヘッダー"
    ws1.append(["売上レポート"])
    ws1.append(["作成日: 2026/4/14"])
    ws1.append([""])  # 空行
    ws1.append(["名前", "部署", "売上"])  # 4行目がヘッダー
    ws1.append(["田中", "営業", 100000])
    ws1.append(["佐藤", "経理", 80000])
    ws1.append(["田中", "営業", 100000])  # 重複
    ws1.append(["鈴木", "開発", 120000])

    # シート2: ヘッダーなし（純粋なデータのみ）
    ws2 = wb.create_sheet("ヘッダーなし")
    ws2.append(["田中", "東京", 30])
    ws2.append(["佐藤", "大阪", 25])
    ws2.append(["田中", "東京", 30])  # 重複
    ws2.append(["鈴木", "名古屋", 40])
    ws2.append(["佐藤", "大阪", 25])  # 重複

    # シート3: セル結合されたタイトルの下にデータ
    ws3 = wb.create_sheet("結合タイトル")
    ws3.merge_cells("A1:C1")
    ws3["A1"] = "2026年度 社員名簿"
    ws3.merge_cells("A2:C2")
    ws3["A2"] = "人事部作成"
    ws3.append([])  # 3行目空
    # 4行目がヘッダー（merge後なのでappendで追加）
    ws3["A4"] = "社員番号"
    ws3["B4"] = "氏名"
    ws3["C4"] = "部署"
    ws3["A5"] = 1001
    ws3["B5"] = "田中太郎"
    ws3["C5"] = "営業部"
    ws3["A6"] = 1002
    ws3["B6"] = "佐藤花子"
    ws3["C6"] = "経理部"
    ws3["A7"] = 1001
    ws3["B7"] = "田中太郎"
    ws3["C7"] = "営業部"  # 重複
    ws3["A8"] = 1003
    ws3["B8"] = "鈴木一郎"
    ws3["C8"] = "開発部"

    # シート4: 途中に空行が挟まるデータ
    ws4 = wb.create_sheet("途中空行")
    ws4.append(["ID", "商品", "価格"])
    ws4.append([1, "りんご", 100])
    ws4.append([2, "みかん", 80])
    ws4.append([])  # 空行
    ws4.append([3, "りんご", 100])  # りんごの重複（空行の後）
    ws4.append([4, "バナナ", 120])
    ws4.append([])  # 空行
    ws4.append([5, "みかん", 80])  # みかんの重複（空行の後）

    wb.save(OUT_DIR / "test6_ヘッダー位置.xlsx")
    print("  test6_ヘッダー位置.xlsx 作成完了")


if __name__ == "__main__":
    print("テスト用Excelファイルを生成中...")
    create_test1()
    create_test2()
    create_test3()
    create_test4()
    create_test5()
    create_test6()
    print(f"完了！ {OUT_DIR} に6ファイル生成しました。")
